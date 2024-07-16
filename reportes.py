import requests
import json
import getpass
import openpyxl
from datetime import datetime
import time

user = input("Por favor ingresa el username: ")
pwd = getpass.getpass("Ahora ingresa la contraseña: ")


url = "https://api.heynowbots.com/api/login"

payload = json.dumps({
  "name": user,
  "password": pwd
})

headers = {
  'Content-Type': 'application/json'
}

# Loggeo en API Hey
response = requests.request("POST", url, headers=headers, data=payload)

response = response.json()
if (response.get('success') == True):
    token = response.get('token')
    
    # Fecha desde la que va a traer reportes
    desde = input("Ingrese la fecha inicial del reporte (aaaa-mm-dd): ")
    hasta = input("Ingrese la fecha final del reporte (aaaa-mm-dd): ")

    flag = True
    count = 0
    reportes = []
    try:
      while (flag):

        if (count == 0):
          url = f'https://api.heynowbots.com:443/api/report/session?begin={desde}&end={hasta}&pageSize=500'
        elif (count > 0):
          url = f'https://api.heynowbots.com:443/api/report/{idScroll}'

        headers = {'Authorization': f'Bearer {token}'}

        response = requests.request("GET", url, headers=headers, data=payload)
        response = response.json()

        length = response['scroll']['length'] - 1 


        """ for item in response['data']:
        keys_to_delete = ['contact', 'chat', 'queryData']
        for key in keys_to_delete:
          if key in item['_source']:
            del item['_source'][key]  """   


        # Mapeo de claves
        mapeo_claves = {
            'bot': ('Canal', lambda x: x['name']),
            'contactFirstName': ('Contacto', lambda x: x),
            'clientId': ('id Contacto', lambda x: x),
            'beginSession': ('Inicio', lambda x: x),
            'endSession': ('Fin', lambda x: x),
            'sessionLength': ('Duracion', lambda x: x),
            'incomingMessages': ('Mensajes entrantes', lambda x: x),
            'outcomingMessages': ('Mensajes salientes', lambda x: x),
            'totalMessages': ('Total de mensajes', lambda x: x),
          }
        for item in response['data']:
          reporte = {}
          # Iterar sobre las claves y extraer los valores
          for key in item['_source']:
            if key in mapeo_claves:
              nombre_reporte, transformacion = mapeo_claves[key]
              reporte[nombre_reporte] = transformacion(item['_source'][key])

          # Verificar si existe startPanelDate
          if item['_source']['startPanelDate'] is not None:
            reporte['Panel'] = 'TRUE'
            reporte['Habilidad'] = item['_source']['abilityList']
            reporte['Agentes'] = item['_source']['agentList']
            reporte['Inicio del panel'] = item['_source']['startPanelDate']
            reporte['Primer contacto'] = item['_source']['firstAgentContact']

            # Formatear tiempo de espera
            milisegundos = int(item['_source']['waitingTime'])
            segundos = milisegundos // 1000
            horas = segundos // 3600
            minutos = (segundos % 3600) // 60
            segundos = segundos % 60
            reporte['Tiempo de espera'] = '{:02d}:{:02d}:{:02d}'.format(horas, minutos, segundos)
          else:
            reporte['Panel'] = 'FALSE'
            reporte['Habilidad'] = ""
            reporte['Agentes'] = ""
            reporte['Inicio del panel'] = ""
            reporte['Primer contacto'] = ""
            reporte['Tiempo de espera'] = ""

          # Verificar si la sesión fue abandonada
          if 'abandoned' in item['_source'] and item['_source']['abandoned']:
            reporte['Sesion abandonada'] = 'TRUE'
          else:
            reporte['Sesion abandonada'] = 'FALSE'
          reportes.append(reporte)
    
        if (response['scroll']['length'] == 0):
          print('Finaliza ejecucion, length = 0')
          flag = False
          break
        else:
          print('Continua la ejecucion...')
          idScroll = response['scroll']['id']
          print(f'Ejecucion: {count}')
          count += 1
          time.sleep(2)
          flag = True


      # Crea el libro y la hoja de Excel
      libro = openpyxl.Workbook()
      # Por defecto .activate selecciona la primer hoja de trabajo
      hoja = libro.active

      # Encabezados de las columnas
      encabezados = ['Canal', 'Contacto', 'id Contacto', 'Inicio', 'Fin', 'Duracion',
                'Mensajes entrantes', 'Mensajes salientes', 'Total de mensajes',
                'Panel', 'Habilidad', 'Agentes', 'Inicio del panel',
                'Primer contacto', 'Tiempo de espera', 'Sesion abandonada']
    
      # Escribir encabezados en la primera fila. Start = 1 indica que debe posicionarse en la primer celda
      for col, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=col, value=encabezado)

      # Convertir las fechas de inicio a objetos datetime
      for reporte in reportes:
        reporte['Inicio'] = datetime.strptime(reporte['Inicio'], '%Y-%m-%dT%H:%M:%S.%fZ')

      # Ordenar los reportes por la fecha de inicio
      reportes_ordenados = sorted(reportes, key=lambda x: x['Inicio'])

      # Escribir datos en el resto de las filas. Inicia en la segunda fila(start=2) y va obteniendo los valores del objeto por la clave del encabezado (start=1)
      for fila, reporte in enumerate(reportes_ordenados, start=2):
        for col, valor in enumerate(encabezados, start=1):
          hoja.cell(row=fila, column=col, value=reporte.get(valor, ""))

      # Guardar el archivo Excel
      libro.save("reporte.xlsx")
    except requests.exceptions.RequestException as e:
      print("Error al hacer la solicitud:", e)
    except json.JSONDecodeError as e:
      print("Error al decodificar JSON:", e)
else:
    print('Error al loguearse')
    print(response)